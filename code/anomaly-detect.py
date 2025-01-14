import pandas as pd
import numpy as np
from sklearn.neighbors import LocalOutlierFactor
import openpyxl
from openpyxl.styles import PatternFill
import os
import constants

def detect_outliers_lof(data, contamination=0.1):
    """
    使用LOF（局部离群因子）检测异常值
    """
    lof = LocalOutlierFactor(n_neighbors=20, contamination=contamination)
    outliers = lof.fit_predict(data)
    return outliers == -1  # LOF返回-1表示异常值

def detect_and_handle_outliers(file_name, output_file, action='clear', contamination=0.1):
    """
    检测并处理Excel文件中的异常值，并将结果保存到新文件
    
    参数:
    file_name (str): 原始Excel文件名（支持相对路径或绝对路径）
    output_file (str): 输出文件名（支持相对路径或绝对路径）
    action (str): 'highlight' 或 'clear'，决定是标红还是清空异常值
    contamination (float): LOF的异常值比例，默认为0.1
    """
    # 将路径转换为绝对路径
    file_path = os.path.abspath(file_name)
    output_path = os.path.abspath(output_file)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在，请检查路径")
        return
    
    # 读取Excel文件
    df = pd.read_excel(file_path, sheet_name='基线阴性插补前')
    
    # 复制数据以保留原始数据
    df_cleaned = df.copy()
    
    # 选择数值型列，并排除不需要处理的列
    numeric_cols = df.select_dtypes(include=['number']).columns
    numeric_cols = [col for col in numeric_cols if col not in constants.id_cols and col not in constants.binary_cols]
    
    # 根据action参数决定操作
    if action == 'highlight':
        # 创建一个新的Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '基线阴性插补前'
        
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # 写入数据
        for r_idx, row in enumerate(df.values, start=2):  # 从第2行开始写入数据
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 定义红色填充
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # 遍历数值型列
        for col in numeric_cols:
            try:
                data = df[[col]].dropna()  # LOF需要二维数据
                outliers = detect_outliers_lof(data, contamination)
                # 获取异常值的位置
                outlier_indices = df.index[outliers]
                # 在新文件中标红异常值
                for idx in outlier_indices:
                    cell = ws.cell(row=idx+2, column=df.columns.get_loc(col)+1)  # +2因为Excel行从1开始，列从1开始
                    cell.fill = red_fill
            except:
                continue  # 跳过无法处理的列
        # 保存新文件
        wb.save(output_path)
        print(f"异常值已标红，结果保存到 {output_path}")
    
    elif action == 'clear':
        # 遍历数值型列
        for col in numeric_cols:
            try:
                data = df_cleaned[[col]].dropna()  # LOF需要二维数据
                outliers = detect_outliers_lof(data, contamination)
                # 将异常值设为空字符串
                df_cleaned.loc[outliers, col] = ""
            except:
                continue  # 跳过无法处理的列
        # 保存清洗后的数据到新文件
        df_cleaned.to_excel(output_path, index=False, header=True)  # 确保表头写入
        print(f"异常值已清空，结果保存到 {output_path}")
    
    else:
        print("无效的操作参数")

# 示例：标红异常值，使用LOF方法
detect_and_handle_outliers(
    file_name='input/negative-baseline.xlsx',  # 支持相对路径或绝对路径
    output_file='temp/highlighted_outliers.xlsx',  # 支持相对路径或绝对路径
    action='highlight',  # 标红异常值
    contamination=0.1  # 异常值比例
)

# 或者：清空异常值，使用LOF方法
# detect_and_handle_outliers(
#     file_name='negative-baseline.xlsx',
#     output_file='cleaned_outliers.xlsx',
#     action='clear',  # 清空异常值
#     contamination=0.1  # 异常值比例
# )