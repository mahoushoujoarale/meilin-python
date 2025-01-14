import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
import os
from constants import binary_cols, id_cols, input_files, sheet_names
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def random_forest_impute_binary(df, target_column):
    """
    对二分类数据进行随机森林插补
    
    参数:
    df (pd.DataFrame): 原始数据
    target_column (str): 需要插补的二分类列名
    
    返回:
    pd.DataFrame: 插补后的数据
    """
    # 分离有缺失值和无缺失值的数据
    missing_data = df[df[target_column].isnull()]
    non_missing_data = df[~df[target_column].isnull()]
    
    # 如果没有缺失值，直接返回
    if missing_data.empty:
        print(f"列 {target_column} 没有缺失值，无需插补")
        return df
    
    # 准备特征和目标变量
    X = non_missing_data.drop(columns=[target_column])
    y = non_missing_data[target_column]
    
    # 将分类变量转换为数值变量
    X = pd.get_dummies(X)
    
    # 训练随机森林模型
    model = RandomForestClassifier(n_estimators=100, random_state=42)
    model.fit(X, y)
    
    # 对缺失值进行预测
    X_missing = missing_data.drop(columns=[target_column])
    X_missing = pd.get_dummies(X_missing)
    
    # 确保特征列一致
    X_missing = X_missing.reindex(columns=X.columns, fill_value=0)
    
    predicted_values = model.predict(X_missing)
    
    # 填补缺失值
    df.loc[df[target_column].isnull(), target_column] = predicted_values
    print(f"二分类列 {target_column} 插补完成")
    
    return df

def multiple_impute_continuous(df):
    """
    对连续数据进行多重插补
    
    参数:
    df (pd.DataFrame): 原始数据
    
    返回:
    pd.DataFrame: 插补后的数据
    """
    # 选择连续型列
    numeric_cols = df.select_dtypes(include=['number']).columns.difference(id_cols).difference(binary_cols)
    
    # 如果没有连续型列，直接返回
    if len(numeric_cols) == 0:
        print("没有连续型列，无需插补")
        return df
    
    # 使用多重插补填补缺失值
    imputer = IterativeImputer(max_iter=10, random_state=42)
    df[numeric_cols] = imputer.fit_transform(df[numeric_cols])
    print("连续数据插补完成")
    
    return df

def highlight_imputed_cells(file_path, df, missing_indices, sheet_name):
    """
    将插补后的单元格标红
    
    参数:
    file_path (str): Excel文件路径
    df (pd.DataFrame): 插补后的数据
    missing_indices (dict): 缺失值的索引字典，键为列名，值为缺失值的索引列表
    sheet_name (str): 工作表名
    """
    # 加载Excel文件
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        print(f"工作表 {sheet_name} 不存在")
        return
    
    # 定义红色填充
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    # 遍历缺失值的索引字典
    for col, indices in missing_indices.items():
        col_idx = df.columns.get_loc(col) + 1  # Excel列索引从1开始
        for row_idx in indices:
            cell = sheet.cell(row=row_idx + 2, column=col_idx)  # Excel行索引从1开始，且第一行为标题
            cell.fill = red_fill
    
    # 保存文件
    book.save(file_path)
    print(f"插补后的单元格已标红，文件保存到 {file_path}")

def impute_data(file_name, sheet_name):
    input_file = "input/" + file_name
    output_file = "temp/imputed-" + file_name
    """
    对数据进行插补（二分类和连续数据）
    
    参数:
    file_name (str): 原始Excel文件名（支持相对路径或绝对路径）
    sheet_name (str): 工作表名
    """
    # 将路径转换为绝对路径
    file_path = os.path.abspath(input_file)
    output_path = os.path.abspath(output_file)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在，请检查路径")
        return
    
    # 读取Excel文件
    df = pd.read_excel(file_path, sheet_name)
    
    # 记录缺失值的索引
    missing_indices = {}
    
    # 对二分类数据进行插补
    if binary_cols:
        for col in binary_cols:
            if col in df.columns:
                missing_indices[col] = df[df[col].isnull()].index.tolist()
                df = random_forest_impute_binary(df, col)
            else:
                print(f"列 {col} 不存在，跳过")
    
    # 对连续数据进行插补
    numeric_cols = df.select_dtypes(include=['number']).columns.difference(id_cols).difference(binary_cols)
    for col in numeric_cols:
        missing_indices[col] = df[df[col].isnull()].index.tolist()
    df = multiple_impute_continuous(df)
    
    # 保存插补后的数据
    df.to_excel(output_path, index=False, sheet_name=sheet_name)
    print(f"插补完成，结果保存到 {output_path}")
    
    # 将插补后的单元格标红
    highlight_imputed_cells(output_path, df, missing_indices, sheet_name)

for file, sheet in zip(input_files, sheet_names):
    impute_data(file, sheet)